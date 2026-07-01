#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
HTML简报生成脚本
基于标引后的Excel数据和内容配置，生成美观的HTML简报
"""
import sys
import os
import base64
import re
import pandas as pd
from openpyxl import load_workbook
import importlib.util
import urllib.request


def download_hero_image():
    """下载Hero背景图片并转换为Base64"""
    url = 'https://images.unsplash.com/photo-1495474472287-4d71bcdd2085?w=1200&q=80'
    try:
        print("正在下载背景图片...")
        with urllib.request.urlopen(url, timeout=10) as response:
            img_data = response.read()
        b64_data = base64.b64encode(img_data).decode('utf-8')
        print(f"背景图片下载成功 ({len(img_data) / 1024:.1f} KB)")
        return f"data:image/jpeg;base64,{b64_data}"
    except Exception as e:
        print(f"背景图片下载失败: {e}，将使用纯渐变背景")
        return None


def load_content_config(tech_topic):
    """动态加载技术主题的内容配置文件"""
    config_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "config")
    config_file = os.path.join(config_dir, f"{tech_topic}_content.py")

    if not os.path.exists(config_file):
        print(f"警告：内容配置文件不存在: {config_file}")
        print(f"将使用默认配置生成简报")
        return None

    # 动态导入配置模块
    spec = importlib.util.spec_from_file_location(f"{tech_topic}_content", config_file)
    config = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(config)

    return config


def load_keywords_config(tech_topic):
    """加载关键词配置（用于获取公司简称映射）"""
    config_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "config")
    config_file = os.path.join(config_dir, f"{tech_topic}_keywords.py")

    if not os.path.exists(config_file):
        return None

    spec = importlib.util.spec_from_file_location(f"{tech_topic}_keywords", config_file)
    config = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(config)

    return config


def extract_images_from_excel(xlsx_path, high_rel_df):
    """从Excel中提取摘要附图"""
    wb = load_workbook(xlsx_path)
    ws = wb.active

    # 提取图片（C列，列索引为2）
    row_to_b64 = {}
    for img in ws._images:
        anchor = img.anchor
        if hasattr(anchor, "_from") and anchor._from.col == 2:  # C列（0-indexed，所以2代表第3列）
            row = anchor._from.row + 1
            try:
                b64 = base64.b64encode(img._data()).decode("utf-8")
                row_to_b64[row] = b64
            except:
                pass

    # 提取超链接（B列是公开号）
    row_to_url = {}
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row, 2)  # B列是公开(公告)号
        if cell.hyperlink:
            row_to_url[row] = cell.hyperlink.target

    # 映射到专利号 - 通过专利号匹配而不是行号
    pn_to_b64 = {}
    pn_to_url = {}

    # 先建立Excel中专利号到行号的映射
    excel_pn_to_row = {}
    for row in range(2, ws.max_row + 1):
        pn = ws.cell(row, 2).value  # B列公开号
        if pn:
            excel_pn_to_row[str(pn)] = row

    # 然后根据high_rel_df中的专利号查找对应的图片和链接
    for i, row in high_rel_df.iterrows():
        pn = str(row["公开(公告)号"])
        if pn in excel_pn_to_row:
            excel_row = excel_pn_to_row[pn]
            if excel_row in row_to_b64:
                pn_to_b64[pn] = row_to_b64[excel_row]
            if excel_row in row_to_url:
                pn_to_url[pn] = row_to_url[excel_row]

    print(f"图片匹配: {len(pn_to_b64)}/{len(high_rel_df)}")
    print(f"链接匹配: {len(pn_to_url)}/{len(high_rel_df)}")

    return pn_to_b64, pn_to_url


def classify_patents_by_keywords(high_rel_df, tech_cats, keywords_config):
    """根据关键词自动分类专利，并为每个分类挑选相关度最高的Top20专利"""
    # 获取搜索字段
    search_cols = getattr(keywords_config, 'SEARCH_COLS', [
        "标题", "Patsnap专利标题", "技术问题", "技术手段", "技术功效"
    ]) if keywords_config else ["标题", "Patsnap专利标题"]

    # 定义六大技术分类的关键词（如果配置中没有提供）
    default_tech_keywords = {
        '研磨精度与均匀度技术': ['研磨', '磨豆', 'grind', 'burr', '刀盘', '颗粒', '磨盘', '研磨机', 'grinder', '磨粉'],
        '加热温控技术': ['加热', '温度', '温控', 'heat', 'temperature', 'thermal', '锅炉', 'boiler', '恒温', '加热器', 'heater'],
        '高压压力动态控制技术': ['压力', '萃取压力', 'pressure', '泵', 'pump', '预浸', 'pre-infusion', '变压', '压力控制'],
        '冲煮头均匀布水萃取技术': ['冲煮', '布水', '萃取', 'brew', 'extraction', 'shower', '淋浴网', '冲泡', '萃取装置'],
        '蒸汽奶泡绵密技术': ['蒸汽', '奶泡', 'steam', 'milk', 'froth', 'foam', '打奶', 'frothing', '奶泡装置', '发泡'],
        '智能电控与水路闭环技术': ['智能', '控制', '传感', 'intelligent', 'control', 'sensor', 'IoT', 'app', '水路', 'flow', '闭环', '自动']
    }

    # 为每个技术分类匹配专利并计算相关度得分
    cat_to_patents = {}
    for catname, catcount, catsummary, cat_patent_list in tech_cats:
        keywords = default_tech_keywords.get(catname, [])
        patent_scores = []

        # 计算所有专利的相关度得分
        for idx, row in high_rel_df.iterrows():
            text = ' '.join(str(row[col]) for col in search_cols if col in row.index and pd.notna(row[col])).lower()

            # 计算匹配的关键词数量作为相关度得分
            score = sum(1 for kw in keywords if kw.lower() in text)

            if score > 0:
                patent_scores.append((str(row['公开(公告)号']), score))

        # 按得分降序排序
        patent_scores.sort(key=lambda x: x[1], reverse=True)

        # 如果配置中提供了代表专利列表，优先保留这些专利，然后补充到20件
        if cat_patent_list and len(cat_patent_list) > 0:
            # 先添加配置中的代表专利
            selected_patents = list(cat_patent_list)
            # 从得分排序中补充，直到达到20件
            for pn, score in patent_scores:
                if pn not in selected_patents and len(selected_patents) < 20:
                    selected_patents.append(pn)
                if len(selected_patents) >= 20:
                    break
            cat_to_patents[catname] = selected_patents
        else:
            # 直接取Top20
            top_patents = [pn for pn, score in patent_scores[:20]]
            cat_to_patents[catname] = top_patents

    return cat_to_patents


def short_name(full, company_short_map):
    """生成公司简称"""
    for k, v in company_short_map.items():
        if k in str(full):
            return v
    name = re.sub(r'(股份有限公司|有限责任公司|有限公司|科技|集团|能源|建材|新能源)', "", str(full))
    return name[:10] if len(name) > 10 else name


def status_tag(s):
    """生成法律状态标签"""
    s = str(s)
    if s == '有效':
        return '<span class="tag tag-valid">有效</span>'
    if s == '审中':
        return '<span class="tag tag-pending">审中</span>'
    if "PCT" in s:
        return '<span class="tag tag-pct">PCT指定期内</span>'
    return f'<span class="tag tag-other">{s}</span>'


def pn_link(pn, pn_to_url):
    """将专利号包装为可点击的链接"""
    url = pn_to_url.get(pn, '#')
    if url == '#':
        return pn
    return f'<a href="{url}" target="_blank" class="patent-link">{pn}</a>'


def img_html(pn, pn_to_b64, cls="patent-img"):
    """生成图片HTML"""
    if pn in pn_to_b64:
        return f'<img class="{cls}" src="data:image/png;base64,{pn_to_b64[pn]}" alt="摘要附图" loading="lazy">'
    return f'<div class="{cls} no-img">暂无附图</div>'


def get_css(hero_bg_url=None):
    """返回CSS样式"""
    # 构建Hero背景样式
    if hero_bg_url:
        hero_bg = f"background:linear-gradient(135deg,rgba(26,26,46,0.85) 0%,rgba(22,33,62,0.80) 60%,rgba(15,52,96,0.75) 100%),url({hero_bg_url}) center/cover;"
    else:
        hero_bg = "background:linear-gradient(135deg,rgba(26,26,46,0.85) 0%,rgba(22,33,62,0.80) 60%,rgba(15,52,96,0.75) 100%);"

    css_a = [
        '* { box-sizing: border-box; margin: 0; padding: 0; }',
        ':root { --orange: #FF6B35; --orange-light: #FF8C42; --orange-pale: #FFF0E8; --dark: #1A1A2E; --dark2: #16213E; --gray: #F5F5F7; --gray2: #E8E8EC; --text: #1D1D1F; --text2: #6E6E73; --white: #FFFFFF; --radius: 12px; --shadow: 0 4px 20px rgba(0,0,0,0.08); --shadow-hover: 0 8px 32px rgba(255,107,53,0.18); }',
        'body { font-family: -apple-system,BlinkMacSystemFont,PingFang SC,Microsoft YaHei,sans-serif; background:var(--gray); color:var(--text); line-height:1.6; }',
        'a{color:var(--orange);text-decoration:none;} a:hover{text-decoration:underline;}',
        '.patent-link{color:var(--orange);font-weight:600;transition:color 0.2s;} .patent-link:hover{color:var(--orange-light);text-decoration:underline;}',
        '.nav{position:sticky;top:0;z-index:100;background:var(--dark);padding:0 40px;display:flex;align-items:center;gap:32px;height:56px;box-shadow:0 2px 12px rgba(0,0,0,0.3);}',
        '.nav-logo{color:var(--orange);font-weight:700;font-size:16px;white-space:nowrap;}',
        '.nav-item{position:relative;} .nav-item>a{color:rgba(255,255,255,0.75);font-size:14px;transition:color 0.2s;white-space:nowrap;cursor:pointer;} .nav-item>a:hover{color:var(--orange);text-decoration:none;}',
        '.nav-dropdown{display:none;position:absolute;top:100%;left:0;background:var(--dark2);min-width:200px;padding:12px 0;border-radius:0 0 8px 8px;box-shadow:0 4px 12px rgba(0,0,0,0.3);margin-top:0;} .nav-item:hover .nav-dropdown{display:block;}',
        '.nav-dropdown a{display:block;color:rgba(255,255,255,0.75);padding:8px 20px;font-size:13px;transition:all 0.2s;} .nav-dropdown a:hover{background:rgba(255,107,53,0.1);color:var(--orange);text-decoration:none;}',
        f'.hero{{position:relative;{hero_bg}color:var(--white);padding:80px 40px 64px;min-height:400px;display:flex;align-items:center;}}',
        '.hero-content{position:relative;z-index:2;max-width:800px;}',
        '.hero-badge{display:inline-block;background:var(--orange);color:white;font-size:12px;font-weight:600;padding:4px 12px;border-radius:20px;margin-bottom:16px;}',
        '.hero h1{font-size:36px;font-weight:700;line-height:1.3;margin-bottom:12px;text-shadow:0 2px 8px rgba(0,0,0,0.3);} .hero h1 span{color:var(--orange-light);}',
        '.hero-meta{color:rgba(255,255,255,0.8);font-size:14px;text-shadow:0 1px 4px rgba(0,0,0,0.3);} .hero-stats{display:flex;gap:40px;margin-top:40px;flex-wrap:wrap;}',
        '.stat{text-align:center;} .stat-num{font-size:36px;font-weight:700;color:var(--orange-light);line-height:1;text-shadow:0 2px 6px rgba(0,0,0,0.3);} .stat-label{font-size:13px;color:rgba(255,255,255,0.8);margin-top:4px;}',
        '.section{padding:56px 40px;} .section-white{background:var(--white);}',
        '.section-header{display:flex;align-items:center;gap:16px;margin-bottom:32px;}',
        '.section-num{width:40px;height:40px;background:var(--orange);color:white;border-radius:10px;display:flex;align-items:center;justify-content:center;font-weight:700;font-size:18px;flex-shrink:0;}',
        '.section-title{font-size:24px;font-weight:700;color:var(--dark);} .section-subtitle{font-size:14px;color:var(--text2);margin-top:2px;}',
        '.news-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(320px,1fr));gap:20px;}',
        '.news-card{background:var(--white);border-radius:var(--radius);padding:24px;box-shadow:var(--shadow);border-left:4px solid var(--orange);transition:box-shadow 0.2s,transform 0.2s;}',
        '.news-card:hover{box-shadow:var(--shadow-hover);transform:translateY(-2px);}',
        '.news-meta{display:flex;gap:8px;align-items:center;margin-bottom:10px;flex-wrap:wrap;}',
        '.news-source{font-size:12px;color:var(--orange);font-weight:600;background:var(--orange-pale);padding:2px 8px;border-radius:4px;}',
        '.news-date{font-size:12px;color:var(--text2);}',
        '.news-title{font-size:15px;font-weight:600;color:var(--dark);margin-bottom:8px;line-height:1.4;}',
        '.news-summary{font-size:13px;color:var(--text2);line-height:1.6;margin-bottom:12px;}',
    ]

    css_b = [
        '.company-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(380px,1fr));gap:20px;}',
        '.company-card{background:var(--white);border-radius:var(--radius);padding:24px;box-shadow:var(--shadow);transition:box-shadow 0.2s,transform 0.2s;}',
        '.company-card:hover{box-shadow:var(--shadow-hover);transform:translateY(-2px);}',
        '.company-header{display:flex;align-items:center;gap:12px;margin-bottom:14px;}',
        '.company-icon{width:44px;height:44px;background:linear-gradient(135deg,var(--orange),var(--orange-light));border-radius:10px;display:flex;align-items:center;justify-content:center;color:white;font-weight:700;font-size:16px;flex-shrink:0;}',
        '.company-name{font-size:16px;font-weight:700;color:var(--dark);} .company-tag{font-size:11px;color:var(--text2);background:var(--gray2);padding:2px 8px;border-radius:4px;margin-top:2px;display:inline-block;}',
        '.company-count{margin-left:auto;text-align:right;} .company-count-num{font-size:22px;font-weight:700;color:var(--orange);line-height:1;} .company-count-label{font-size:11px;color:var(--text2);}',
        '.company-summary{font-size:13px;color:var(--text2);line-height:1.7;margin-bottom:14px;}',
        '.company-patents{display:flex;flex-direction:column;gap:6px;} .company-patent-item{font-size:12px;background:var(--gray);padding:6px 10px;border-radius:6px;} .company-patent-pn{color:var(--orange);font-weight:600;margin-right:6px;}',
        '.tech-cat{margin-bottom:48px;}',
        '.tech-cat-header{background:linear-gradient(90deg,var(--dark) 0%,var(--dark2) 100%);color:white;padding:20px 28px;border-radius:var(--radius) var(--radius) 0 0;display:flex;align-items:center;gap:16px;}',
        '.tech-cat-name{font-size:18px;font-weight:700;} .tech-cat-badge{background:var(--orange);color:white;font-size:12px;font-weight:600;padding:3px 10px;border-radius:12px;}',
        '.tech-cat-body{background:var(--white);border-radius:0 0 var(--radius) var(--radius);padding:24px 28px;box-shadow:var(--shadow);}',
        '.tech-cat-summary{font-size:14px;color:var(--text2);line-height:1.8;margin-bottom:24px;padding-bottom:20px;border-bottom:1px solid var(--gray2);}',
        '.tech-cat-subtitle{font-size:13px;font-weight:600;color:var(--text2);text-transform:uppercase;letter-spacing:0.5px;margin-bottom:16px;}',
        '.patent-grid-sm{display:grid;grid-template-columns:repeat(auto-fill,minmax(260px,1fr));gap:16px;}',
        '.patent-card-sm{background:var(--gray);border-radius:10px;padding:16px;transition:box-shadow 0.2s;} .patent-card-sm:hover{box-shadow:var(--shadow-hover);}',
        '.patent-img-sm{width:100%;height:140px;object-fit:contain;background:white;border-radius:6px;margin-bottom:12px;}',
        '.no-img{display:flex;align-items:center;justify-content:center;color:var(--text2);font-size:12px;background:var(--gray2);}',
        '.patent-pn{font-size:12px;font-weight:700;color:var(--orange);margin-bottom:4px;}',
        '.patent-title-sm{font-size:13px;font-weight:600;color:var(--dark);margin-bottom:8px;line-height:1.4;}',
        '.patent-patsnap{font-size:12px;color:var(--text2);line-height:1.6;margin-bottom:10px;}',
        '.patent-meta{display:flex;flex-wrap:wrap;gap:6px;}',
        '.tag{font-size:11px;padding:2px 8px;border-radius:4px;font-weight:500;}',
        '.tag-valid{background:#E8F5E9;color:#2E7D32;} .tag-pending{background:#E3F2FD;color:#1565C0;} .tag-pct{background:#FFF3E0;color:#E65100;} .tag-other{background:var(--gray2);color:var(--text2);} .tag-cat{background:var(--orange-pale);color:var(--orange);} .tag-company{background:#F3E5F5;color:#6A1B9A;}',
        '.footer{background:var(--dark);color:rgba(255,255,255,0.5);text-align:center;padding:32px 40px;font-size:13px;}',
        '.footer span{color:var(--orange);}',
        '@media(max-width:768px){.nav{padding:0 16px;gap:16px;overflow-x:auto;} .hero{padding:40px 16px;} .hero h1{font-size:22px;} .section{padding:40px 16px;} .news-grid,.company-grid,.patent-grid,.patent-grid-sm{grid-template-columns:1fr;}}',
    ]

    return "\n".join(css_a + css_b)


def build_html(tech_topic, date_range, high_rel_df, pn_to_b64, pn_to_url, config, company_short_map, keywords_config, hero_bg_url=None):
    """构建HTML内容"""
    p = []
    A = p.append

    # 获取配置数据
    news = getattr(config, 'NEWS', []) if config else []
    company_data = getattr(config, 'COMPANY_DATA', []) if config else []
    tech_cats = getattr(config, 'TECH_CATS', []) if config else []

    # 根据关键词自动分类专利
    cat_to_patents = classify_patents_by_keywords(high_rel_df, tech_cats, keywords_config) if tech_cats else {}

    # 统计数据
    total_patents = len(high_rel_df)
    num_companies = len(company_data) if company_data else high_rel_df['[标]当前申请(专利权)人'].nunique()
    num_tech_cats = len(tech_cats) if tech_cats else (high_rel_df['技术分类'].nunique() if '技术分类' in high_rel_df.columns else 0)
    num_news = len(news)

    # HTML头部
    A('<!DOCTYPE html><html lang="zh-CN"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">')
    A(f'<title>{tech_topic}研发简报 {date_range}</title><style>{get_css(hero_bg_url)}</style></head><body>')

    # 导航栏
    A(f'<nav class="nav"><span class="nav-logo">{tech_topic}研发简报</span>')
    if news:
        A('<a href="#news">行业动态</a>')
    if company_data:
        A('<div class="nav-item"><a href="#company">重点公司</a><div class="nav-dropdown">')
        for cname, _, _, _ in company_data:
            company_id = 'company-' + cname.replace(' ', '').replace('(', '').replace(')', '')
            A(f'<a href="#{company_id}">{cname}</a>')
        A('</div></div>')
    if tech_cats:
        A('<div class="nav-item"><a href="#tech">技术分类与专利解读</a><div class="nav-dropdown">')
        for catname, _, _, _ in tech_cats:
            cat_id = 'tech-' + catname
            A(f'<a href="#{cat_id}">{catname}</a>')
        A('</div></div>')
    A('</nav>')

    # Hero区域
    A('<div class="hero"><div class="hero-content"><div class="hero-badge">研发情报</div>')
    A(f'<h1>{tech_topic}行业研发简报<br><span>{date_range}</span></h1>')
    A('<div class="hero-meta">数据来源：Patsnap专利数据库</div>')
    A('<div class="hero-stats">')
    A(f'<div class="stat"><div class="stat-num">{total_patents}</div><div class="stat-label">高相关专利</div></div>')
    if num_companies > 0:
        A(f'<div class="stat"><div class="stat-num">{num_companies}</div><div class="stat-label">重点公司</div></div>')
    if num_tech_cats > 0:
        A(f'<div class="stat"><div class="stat-num">{num_tech_cats}</div><div class="stat-label">技术分类</div></div>')
    if num_news > 0:
        A(f'<div class="stat"><div class="stat-num">{num_news}</div><div class="stat-label">行业动态</div></div>')
    A('</div></div></div>')

    # 行业动态
    if news:
        A('<section class="section" id="news"><div class="section-header"><div class="section-num">1</div>')
        A(f'<div><div class="section-title">{tech_topic}行业动态</div><div class="section-subtitle">重要行业资讯</div></div></div>')
        A('<div class="news-grid">')
        for title, source, date, summary, url in news:
            A(f'<div class="news-card"><div class="news-meta"><span class="news-source">{source}</span><span class="news-date">{date}</span></div>')
            A(f'<div class="news-title">{title}</div><div class="news-summary">{summary}</div>')
            A(f'<a class="news-link" href="{url}" target="_blank">查看原文 →</a></div>')
        A('</div></section>')

    # 重点公司
    if company_data:
        section_num = 2 if news else 1
        A(f'<section class="section section-white" id="company"><div class="section-header"><div class="section-num">{section_num}</div>')
        A('<div><div class="section-title">重点公司技术动向</div><div class="section-subtitle">重点企业本期专利技术进展</div></div></div>')
        A('<div class="company-grid">')
        for cname, ctag, csummary, cpns in company_data:
            icon = cname[0]
            company_id = 'company-' + cname.replace(' ', '').replace('(', '').replace(')', '')
            A(f'<div class="company-card" id="{company_id}"><div class="company-header">')
            A(f'<div class="company-icon">{icon}</div><div><div class="company-name">{cname}</div><span class="company-tag">{ctag}</span></div>')
            A(f'<div class="company-count"><div class="company-count-num">{len(cpns)}</div><div class="company-count-label">本期专利</div></div></div>')
            A(f'<div class="company-summary">{csummary}</div><div class="company-patents">')
            for pn in cpns:
                row = high_rel_df[high_rel_df["公开(公告)号"] == pn]
                ptitle = row["标题"].values[0] if len(row) > 0 else pn
                A(f'<div class="company-patent-item"><span class="company-patent-pn">{pn_link(pn, pn_to_url)}</span>{ptitle}</div>')
            A('</div></div>')
        A('</div></section>')

    # 技术分类
    if tech_cats:
        section_num = 3 if news and company_data else (2 if news or company_data else 1)
        A(f'<section class="section" id="tech"><div class="section-header"><div class="section-num">{section_num}</div>')
        A(f'<div><div class="section-title">技术分类与专利解读</div><div class="section-subtitle">高相关专利按技术方向归类解读</div></div></div>')

        for catname, catcount, catsummary, cat_patent_list in tech_cats:
            # 使用自动分类的结果
            patent_list = cat_to_patents.get(catname, [])
            cat_patents = high_rel_df[high_rel_df["公开(公告)号"].isin(patent_list)]

            cat_id = 'tech-' + catname
            A(f'<div class="tech-cat" id="{cat_id}"><div class="tech-cat-header">')
            A(f'<span class="tech-cat-name">{catname}</span><span class="tech-cat-badge">{len(cat_patents)}条专利</span></div>')
            A(f'<div class="tech-cat-body"><p class="tech-cat-summary">{catsummary}</p>')
            A('<div class="tech-cat-subtitle">代表专利</div><div class="patent-grid-sm">')

            for i, r in cat_patents.iterrows():
                pn = str(r["公开(公告)号"])
                ptitle = str(r["标题"])
                applicant = str(r.get("[标]当前申请(专利权)人", "N/A"))
                status = str(r.get("简单法律状态", "N/A"))
                patsnap = str(r.get("Patsnap专利标题", "-"))
                if patsnap == "nan":
                    patsnap = "-"

                A('<div class="patent-card-sm">')
                A(img_html(pn, pn_to_b64, "patent-img-sm"))
                A(f'<div class="patent-pn">{pn_link(pn, pn_to_url)}</div>')
                A(f'<div class="patent-title-sm">{ptitle}</div>')
                A(f'<div class="patent-patsnap">{patsnap}</div>')
                A(f'<div class="patent-meta">{status_tag(status)}<span class="tag tag-company">{short_name(applicant, company_short_map)}</span><span class="tag tag-cat">{catname}</span></div>')
                A('</div>')

            A('</div></div></div>')
        A('</section>')

    # 页脚
    A(f'<footer class="footer">{tech_topic}研发简报 <span>{date_range}</span> | 数据来源：Patsnap专利数据库</footer>')
    A('</body></html>')

    return "".join(p)


def main():
    if len(sys.argv) < 5:
        print("用法: python generate_report.py <标引Excel> <输出HTML> <技术主题> <日期范围>")
        sys.exit(1)

    tagged_xlsx = sys.argv[1]
    output_html = sys.argv[2]
    tech_topic = sys.argv[3]
    date_range = sys.argv[4]

    # 加载配置
    content_config = load_content_config(tech_topic)
    keywords_config = load_keywords_config(tech_topic)

    # 获取公司简称映射
    company_short_map = getattr(keywords_config, 'COMPANY_SHORT', {}) if keywords_config else {}

    # 读取标引后的Excel
    df = pd.read_excel(tagged_xlsx, engine="openpyxl")

    # 筛选高相关专利
    if "相关性标引" in df.columns:
        high_rel = df[df["相关性标引"] == "高相关"].copy().reset_index(drop=True)
    elif "技术相关性" in df.columns:
        high_rel = df[df["技术相关性"] == "高相关"].copy().reset_index(drop=True)
    else:
        print("警告：未找到相关性标引列，将使用所有专利")
        high_rel = df.copy().reset_index(drop=True)

    print(f"高相关专利数: {len(high_rel)}")

    # 提取图片和链接
    pn_to_b64, pn_to_url = extract_images_from_excel(tagged_xlsx, high_rel)

    # 下载并转换Hero背景图片
    hero_bg_url = download_hero_image()

    # 生成HTML
    html = build_html(tech_topic, date_range, high_rel, pn_to_b64, pn_to_url, content_config, company_short_map, keywords_config, hero_bg_url)

    # 写入文件
    with open(output_html, "w", encoding="utf-8") as f:
        f.write(html)

    print(f"\n生成完成: {output_html}")
    print(f"文件大小: {len(html) // 1024} KB")


if __name__ == "__main__":
    main()
