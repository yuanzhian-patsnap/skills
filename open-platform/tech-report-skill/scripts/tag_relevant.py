#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
专利相关性标引脚本
根据关键词配置文件，自动标记高相关专利
"""
import sys
import os
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font
import importlib.util

sys.stdout.reconfigure(encoding="utf-8")


def load_keywords_config(tech_topic):
    """动态加载技术主题的关键词配置文件"""
    config_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "config")
    config_file = os.path.join(config_dir, f"{tech_topic}_keywords.py")

    if not os.path.exists(config_file):
        print(f"错误：配置文件不存在: {config_file}")
        print(f"请先创建 {tech_topic}_keywords.py 配置文件")
        sys.exit(1)

    # 动态导入配置模块
    spec = importlib.util.spec_from_file_location(f"{tech_topic}_keywords", config_file)
    config = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(config)

    return config


def is_high_relevant(row, include_keywords, exclude_keywords, search_cols):
    """判断专利是否高相关"""
    text = " ".join(
        str(row[col]) for col in search_cols if col in row.index and pd.notna(row[col])
    ).lower()

    # 命中排除词 → 不标引
    for kw in exclude_keywords:
        if kw.lower() in text:
            return False

    # 命中包含词 → 高相关
    for kw in include_keywords:
        if kw.lower() in text:
            return True

    return False


def main():
    if len(sys.argv) < 4:
        print("用法: python tag_relevant.py <输入Excel> <输出Excel> <技术主题>")
        sys.exit(1)

    input_file = sys.argv[1]
    output_file = sys.argv[2]
    tech_topic = sys.argv[3]

    # 加载配置
    config = load_keywords_config(tech_topic)
    include_keywords = config.INCLUDE_KEYWORDS
    exclude_keywords = config.EXCLUDE_KEYWORDS

    # 检索字段（优先使用配置文件中的，否则使用默认值）
    search_cols = getattr(config, 'SEARCH_COLS', [
        "标题", "Patsnap专利标题", "技术问题", "技术手段", "技术功效"
    ])

    # 读取Excel
    df = pd.read_excel(input_file, engine="openpyxl")

    # 标引
    df["相关性标引"] = df.apply(
        lambda row: "高相关" if is_high_relevant(row, include_keywords, exclude_keywords, search_cols) else "",
        axis=1
    )

    # 统计
    high_count = (df["相关性标引"] == "高相关").sum()
    print(f"总专利数: {len(df)}")
    print(f"高相关专利数: {high_count}")
    print(f"高相关比例: {high_count/len(df)*100:.1f}%")
    print()
    print("高相关专利列表:")
    for _, r in df[df["相关性标引"] == "高相关"].iterrows():
        pn = r.get("公开(公告)号", "N/A")
        title = r.get("标题", "N/A")
        print(f"  [{pn}] {title}")

    # 直接在原始Excel文件上添加标引列，保留图片和超链接
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active

    # 在最后一列添加"相关性标引"列
    tag_col = ws.max_column + 1
    ws.cell(1, tag_col).value = "相关性标引"

    # 填充标引结果
    for idx, value in enumerate(df["相关性标引"], start=2):
        ws.cell(idx, tag_col).value = value

    # 高亮高相关行
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    red_font = Font(color="CC0000", bold=True)

    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row, tag_col)
        if cell.value == "高相关":
            cell.font = red_font
            for col in range(1, ws.max_column + 1):
                ws.cell(row, col).fill = yellow_fill

    wb.save(output_file)
    print(f"\n已输出: {output_file}")


if __name__ == "__main__":
    main()
